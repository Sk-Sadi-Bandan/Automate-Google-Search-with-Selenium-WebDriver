from lib2to3.pgen2 import driver

import openpyxl
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from datetime import datetime

# Define the path to your Excel file
excel_file_path = r"C:\Users\hp\Downloads\Python_selium-main\Python_selium-main\Excel.xlsx"

# Define a function to get Google search suggestions for a given keyword
def get_google_suggestions(keyword):
    firefox_driver_path = r"C:\Users\hp\Downloads\Compressed\geckodriver-v0.33.0-win64"

    os.environ["PATH"] += os.pathsep + firefox_driver_path
    driver = webdriver.Firefox()
    driver.get("https://www.google.com")
    search_box = driver.find_element(By.NAME, "q")
    search_box.clear()
    search_box.send_keys(keyword)
    
    wait = WebDriverWait(driver, 10)
    wait.until(EC.presence_of_element_located((By.XPATH, "//ul[@role='listbox']/li[@role='presentation']")))

    suggestions = driver.find_elements(By.XPATH, "//ul[@role='listbox']/li[@role='presentation']")
    suggestion_texts = [suggestion.text for suggestion in suggestions]
    
    driver.quit()
    
    if suggestion_texts:
        longest_suggestion = max(suggestion_texts, key=len)
        shortest_suggestion = min(suggestion_texts, key=len)
        return longest_suggestion, shortest_suggestion
    else:
        return None, None


workbook = openpyxl.load_workbook(excel_file_path)

# Get the current day
current_day = datetime.now().strftime("%A")


if current_day in workbook.sheetnames:
    worksheet = workbook[current_day]

   
    longest_suggestions = []
    shortest_suggestions = []

    # Iterate through the rows and get suggestions
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=2, max_col=2):
        keyword = row[0].value
        if keyword and not keyword.isspace():
            long_suggestion, short_suggestion = get_google_suggestions(keyword)
            longest_suggestions.append(long_suggestion)
            shortest_suggestions.append(short_suggestion)

    
    for idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=3, max_col=4)):
        row[0].value = longest_suggestions[idx]
        row[1].value = shortest_suggestions[idx]

    # Save the updated Excel file
    workbook.save(excel_file_path)

   
    workbook.close()


driver.quit()
